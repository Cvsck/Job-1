import json
import logging
from datetime import datetime
from functools import reduce
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

# Настройка логирования
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


def parse_date(date_str: Optional[str]) -> Optional[datetime]:
    """Преобразование строки даты в объект datetime."""
    if date_str is None:
        return None
    try:
        return datetime.strptime(date_str, "%d.%m.%Y")
    except (ValueError, TypeError):
        try:
            return datetime.strptime(date_str, "%d.%m.%Y %H:%M:%S")
        except (ValueError, TypeError):
            return None


def filter_transactions_by_date(transactions: List[Dict[str, Any]], year: int, month: int) -> List[Dict[str, Any]]:
    """Фильтрация транзакций по году и месяцу."""
    return list(
        filter(
            lambda record: record["Дата операции"]
            and isinstance(record["Дата операции"], datetime)
            and record["Дата операции"].year == year
            and record["Дата операции"].month == month
            and record["Сумма операции"] < 0,
            transactions,
        )
    )



def calculate_cashback_by_category(filtered_transactions: List[Dict[str, Any]]) -> Dict[str, float]:
    """Расчет кешбэка по категориям."""
    cashback_rate = 0.01

    cashback_by_category = reduce(
        lambda acc, transaction: {
            **acc,
            transaction["Категория"]: acc.get(transaction["Категория"], 0)
            + abs(transaction["Сумма операции"]) * cashback_rate,
        },
        filtered_transactions,
        {},
    )

    # Сортировка категорий по убыванию кешбэка
    sorted_cashback_by_category = dict(sorted(cashback_by_category.items(), key=lambda item: item[1], reverse=True))

    return {category: round(cashback, 2) for category, cashback in sorted_cashback_by_category.items()}


def analyze_cashback_categories(
    transactions_data: List[Dict[str, Any]], filter_year: int, filter_month: int
) -> Dict[str, float]:
    """Основная функция анализа кешбэка по категориям."""
    transactions_data = list(
        map(
            lambda record: {
                **record,
                "Дата операции": parse_date(record.get("Дата операции") or record.get("Дата платежа")),
            },
            transactions_data,
        )
    )

    filtered_transactions = filter_transactions_by_date(transactions_data, filter_year, filter_month)

    if not filtered_transactions:
        logger.warning(f"Транзакции не найдены за {filter_year}-{filter_month}")
        return {}

    cashback_by_category = calculate_cashback_by_category(filtered_transactions)
    logger.info(f"Завершен анализ категорий кешбэка за {filter_year}-{filter_month}")
    return cashback_by_category


def load_data_from_excel(file_path: str) -> List[Dict[str, Any]]:
    """Загрузка данных из Excel файла."""
    workbook = load_workbook(file_path)
    sheet = workbook.active

    header = [cell.value for cell in sheet[1]]
    data = [dict(zip(header, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]

    logger.info(f"Загружено {len(data)} записей из {file_path}")
    return data


def save_result_to_file(result: Dict[str, float], file_path: str) -> None:
    """Сохранение результата в файл JSON."""
    with open(file_path, "w", encoding="utf-8") as file:
        json.dump(result, file, ensure_ascii=False, indent=4)
    logger.info(f"Результаты сохранены в файл {file_path}")
